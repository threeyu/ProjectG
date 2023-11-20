import os
import argparse
from tkinter import filedialog
import openpyxl
import json
import shutil

parser = argparse.ArgumentParser()
parser.add_argument('-w', help='Set working directory')
parser.add_argument('-c', help='Set config xlsx file')
parser.add_argument('-m', help='Set tool mode. data: convert excel to gamedata. language: generate language excel', choices=['data', 'language'])
parser.add_argument('--all_keys_out_path', help='Set all key json output path')
parser.add_argument('--json', help='Set json output folder')
parser.add_argument('--lang_out_path', help='Set language excel output path')
parser.add_argument('--lang_config', help='Set language config excel file path')

def get_xlsx_config(cfg_fp):
    workbook = openpyxl.load_workbook(cfg_fp)
    table = workbook.active

    config = {
        'tables': []
    }
    print(f'[ExcelTool] Total table count: {table.max_row - 1}')
    for row in table.iter_rows(min_row=2):
        config['tables'].append({
            'filename': str(row[0].value),
            'sheet': str(row[1].value),
            'result': str(row[2].value),
            'mapper': row[3].value if row[3].value is not None else ''
        })
    
    print(f'[ExcelTool] Config: {config}')

    return config

def parse_field(idx, row, defaults, typ):
    if typ == 'string':
        if row[idx].value is None:
            if defaults[idx] is None:
                return ''
            else:
                return str(defaults[idx])
        else:
            return str(row[idx].value)
    elif typ == 'int':
        if row[idx].value is None:
            if defaults[idx] is None:
                return 0
            else:
                return int(defaults[idx])
        else:
            return int(row[idx].value)
    elif typ == 'float':
        if row[idx].value is None:
            if defaults[idx] is None:
                return 0.0
            else:
                return float(defaults[idx])
        else:
            return float(row[idx].value)
    elif typ == 'list':
        if row[idx].value is None:
            if defaults[idx] is None:
                return []
            else:
                return str(defaults[idx]).split(';')
        else:
            return str(row[idx].value).split(';')
    elif typ == 'int_list':
        if row[idx].value is None:
            if defaults[idx] is None:
                return []
            else:
                return list([int(v) for v in str(defaults[idx]).split(';')])
        else:
            return list([int(v) for v in str(row[idx].value).split(';')])
    elif typ == 'float_list':
        if row[idx].value is None:
            if defaults[idx] is None:
                return []
            else:
                return list([float(v) for v in str(defaults[idx]).split(';')])
        else:
            return list([float(v) for v in str(row[idx].value).split(';')])
    elif typ == 'decimal':
        if row[idx].value is None:
            if defaults[idx] is None:
                return ''
            else:
                return str(defaults[idx])
        else:
            return str(row[idx].value)
    elif typ == 'decimal_list':
        if row[idx].value is None:
            if defaults[idx] is None:
                return []
            else:
                return str(defaults[idx]).split(';')
        else:
            return str(row[idx].value).split(';')

def get_loc_key(table_name, column_name, primary_key):
    return f'{table_name}_{column_name}_{primary_key}'

def parse_table(tcfg, work_dir, all_keys):
    # 普通的配置表
    table_fp = os.path.join(work_dir, tcfg['filename'])
    workbook = openpyxl.load_workbook(table_fp)
    table = workbook[tcfg['sheet']]
    rows = list(table.rows)
    table_name = tcfg['result']
    
    keys = list([c.value for c in rows[0] if c.value is not None])
    print(keys)
    # desc = table.row_values(1)
    types = list([c.value for c in rows[2]])
    defaults = list([c.value for c in rows[3]])
    tags = list([c.value for c in rows[4]])

    result = {}
    primary_key_idx = 0

    for idx, tag in enumerate(tags):
        if tag is not None and 'primary' in tag:
            primary_key_idx = idx
    primary_key = keys[primary_key_idx]

    print(f'[ExcelTool] Sheet {tcfg["filename"]} has {table.max_row - 5} rows')
    print(f'[ExcelTool] Primary index is {primary_key_idx}, primary key is {primary_key}')
    for row in table.iter_rows(min_row=6):
        data = {}
        if row[primary_key_idx].value is None:
            continue
        for idx, key in enumerate(keys):
            # 被忽略的列
            if tags[idx] is not None and 'ignore' in tags[idx]:
                continue
            data[key] = parse_field(idx, row, defaults, types[idx])

            if tags[idx] is not None and 'localization' in tags[idx]:
                okey = f'{table_name}_{key}_{row[primary_key_idx].value}'
                all_keys[okey] = row[idx].value
        result[data[primary_key]] = data
    return result

def parse_single_cell(value, typ, default):
    if typ == 'string':
        if value is None:
            if default is None:
                return ''
            else:
                return str(default)
        else:
            return str(value)
    elif typ == 'int':
        if value is None:
            if default is None:
                return 0
            else:
                return int(default)
        else:
            return int(value)
    elif typ == 'float':
        if value is None:
            if default is None:
                return 0.0
            else:
                return float(default)
        else:
            return float(value)
    elif typ == 'list':
        if value is None:
            if default is None:
                return []
            else:
                return str(default).split(';')
        else:
            return str(value).split(';')
    elif typ == 'int_list':
        if value is None:
            if default is None:
                return []
            else:
                return list([int(v) for v in str(default).split(';')])
        else:
            return list([int(v) for v in str(value).split(';')])
    elif typ == 'float_list':
        if value is None:
            if default is None:
                return []
            else:
                return list([float(v) for v in str(default).split(';')])
        else:
            return list([float(v) for v in str(value).split(';')])
    elif typ == 'decimal':
        if value is None:
            if default is None:
                return ''
            else:
                return str(default)
        else:
            return str(value)
    elif typ == 'decimal_list':
        if value is None:
            if default is None:
                return []
            else:
                return str(default).split(';')
        else:
            return str(value).split(';')

def parse_properties_table(tcfg, work_dir, all_keys):
    # 角色属性的配置表，每一个页签是一个角色
    print(f'[ExcelTool] Parsing character properties table')
    table_fp = os.path.join(work_dir, tcfg['filename'])
    workbook = openpyxl.load_workbook(table_fp)

    result = {}

    for sname in workbook.sheetnames:
        print(f'[ExcelTool] Parsing sheet ' + sname)
        sheet = workbook[sname]
        rows = list(sheet.rows)
        keys = list([c.value for c in rows[0] if c.value is not None])
        types = list([c.value for c in rows[2]])
        defaults = list([c.value for c in rows[3]])
        tags = list([c.value for c in rows[4]])

        # 找主键
        primary_key_index = 0
        for idx, tag in enumerate(tags):
            if tag is not None and 'primary' in tag:
                primary_key_index = idx
        primary_key = keys[primary_key_index]

        data = {} # 单个行结果对象

        # 创建key的空列表
        for key in keys:
            data[key] = []

        for row in sheet.iter_rows(min_row=6):
            if row[primary_key_index].value is None:
                continue
            for idx, key in enumerate(keys):
                # 忽略某列
                if tags[idx] is not None and 'ignore' in tags[idx]:
                    continue
                val = parse_single_cell(row[idx].value, types[idx], defaults[idx])
                data[key].append(val)
        
        result[sname] = data

    return result

def lkeys_from_normal_table(tcfg, work_dir):
    table_fp = os.path.join(work_dir, tcfg['filename'])
    workbook = openpyxl.load_workbook(table_fp)
    table = workbook[tcfg['sheet']]
    table_name = tcfg['result']
    rows = list(table.rows)

    keys = list([c.value for c in rows[0]])
    # desc = table.row_values(1)
    types = list([c.value for c in rows[2]])
    defaults = list([c.value for c in rows[3]])
    tags = list([c.value if c.value is not None else '' for c in rows[4]])

    primary_key_idx = 0

    # 找primary key
    for idx, tag in enumerate(tags):
        if tag is not None and 'primary' in tag:
            primary_key_idx = idx
    primary_key = keys[primary_key_idx]

    okeys = {}
    for row in table.iter_rows(min_row=6):
        if row[primary_key_idx].value is None:
            continue
        id = row[primary_key_idx].value
        for idx, field in enumerate(keys):
            # 被忽略的列
            okey = f'{table_name}_{field}_{id}'

            if 'ignore' in tags[idx] or 'localization' not in tags[idx]:
                continue

            # 获取内容
            content = ''
            if row[idx].value is None:
                if defaults[idx] is None:
                    content = ''
                else:
                    content = str(defaults[idx])
            else:
                content = str(row[idx].value)

            okeys[okey] = content
    return okeys

def lkeys_from_string_table(tcfg, work_dir):
    table_fp = os.path.join(work_dir, tcfg['filename'])
    workbook = openpyxl.load_workbook(table_fp)
    table = workbook[tcfg['sheet']]

    key_list = {}
    for row in table.iter_rows(min_row=6):
        key = row[0].value
        value = row[1].value
        note = row[2].value if row[2].value is not None else ''
        key_list[key] = {
            'content': value,
            'note': note
        }

    return key_list

def ensure_empty_dir(folder):
    # 重新创建文件夹
    if os.path.exists(folder):
        shutil.rmtree(folder)
    os.makedirs(folder)

mapper_func_table = {
    'CharacterProperties': parse_properties_table,
}

def convert_excel_to_data(work_dir, json_out_dir, cfg_fp, all_keys_out_path):
    config = get_xlsx_config(cfg_fp)
    ensure_empty_dir(json_out_dir)
    all_keys = {}
    for tcfg in config['tables']:
        print('[ExcelTool] Start to convert excel', tcfg['filename'], 'sheet', tcfg['sheet'], 'result name is', tcfg['result'], 'mapper:', tcfg['mapper'])
        parsed = {}
        mapper_name = tcfg['mapper']
        mapper_func = parse_table
        if mapper_name in mapper_func_table:
            mapper_func = mapper_func_table[mapper_name]
        parsed = mapper_func(tcfg, work_dir, all_keys)

        # 保存json文件
        json_fp = os.path.join(json_out_dir, tcfg['result'] + '.bytes')
        with open(json_fp, 'w', encoding='utf-8') as f:
            f.write(json.dumps(parsed, ensure_ascii=False))
        print('[ExcelTool] Successfully converted excel', tcfg['filename'], 'sheet', tcfg['sheet'], 'to json', json_fp)
        
        print('[ExcelTool] Next\n')
    
    # 是否保存全部key值
    if all_keys_out_path is not None:
        with open(all_keys_out_path, 'w', encoding='utf-8') as f:
            f.write(json.dumps(all_keys, ensure_ascii=False))

class LanguageConfig:
    def __init__(self):
        self.key_config = {}
    
def language_config_from_file(path):
    xlsx = openpyxl.load_workbook(path)
    config = LanguageConfig()

    table = xlsx['config']
    designer_idx = 0
    table_idx = 1
    field_idx = 2
    id_idx = 3
    tag_idx = 4
    note_idx = 5
    for row in table.iter_rows(min_row=1):
        if row[table_idx].value is None:
            continue
        okey = f'{row[table_idx].value}_{row[field_idx].value}_{row[id_idx].value}'
        config.key_config[okey] = {
            'designer': row[designer_idx].value,
            'tag': row[tag_idx].value,
            'note': row[note_idx].value
        }
    
    print(config.key_config)

    return config

def generate_language_excel(work_dir, language_config_path, cfg_fp, lang_out_path):
    config = get_xlsx_config(cfg_fp)
    print('[ExcelTool] Start to generate language excel')

    lconfig = language_config_from_file(language_config_path)

    key_list = {}
    all_okeys = {}
    for tcfg in config['tables']:
        print('[ExcelTool] Generate keys from file ', tcfg['filename'], 'sheet', tcfg['sheet'])
        mapper_name = tcfg['mapper']
        if mapper_name == 'CharacterProperties':
            continue
        elif mapper_name == 'Strings':
            final_keys = lkeys_from_string_table(tcfg, work_dir)
            key_list.update(final_keys)
        else:
            keys = lkeys_from_normal_table(tcfg, work_dir)
            all_okeys.update(keys)

    for okey in all_okeys:
        version = '1'
        designer = 'S'
        tag = ''
        note = ''
        if okey in lconfig.key_config:
            designer = lconfig.key_config[okey]['designer']
            tag = lconfig.key_config[okey]['tag']
            note = lconfig.key_config[okey]['note']
        
        parts = [
            'O',
            version,
            designer,
            okey
        ]
        if len(tag) > 0:
            parts.append(tag)
        final_key = '_'.join(parts)
        key_list[final_key] = {
            'content': all_okeys[okey],
            'note': note
        }

    # 导出到excel
    header = 'key,content,note\n'
    lines = []
    for key, value in key_list.items():
        content = value['content']
        note = value['note']
        lines.append(f'{key},{content},{note}')
    ctx = '\n'.join(lines)
    with open(lang_out_path, 'w') as f:
        f.write(f'{header}{ctx}')

def main():
    args = parser.parse_args()
    print('[ExcelTool] Mode:', args.m)

    work_dir = os.getcwd()
    if args.w:
        work_dir = os.path.abspath(args.w)
    print('[ExcelTool] Working path:', work_dir)

    cfg_fp = os.path.join(work_dir, 'Config.xlsx')
    if args.c:
        cfg_fp = os.path.abspath(args.c)
    print('[ExcelTool] Config absolute path:', cfg_fp)

    lang_cfg_fp = os.path.join(work_dir, 'Localization/LocalizationConfig.xlsx')
    if args.lang_config:
        lang_cfg_fp = os.path.abspath(args.lang_config)
    print('[ExcelTool] Language config path:', lang_cfg_fp)

    json_out_dir = os.path.join(work_dir, 'GameData')
    if args.json:
        json_out_dir = os.path.abspath(args.json)
    print('[ExcelTool] Json output directory:', json_out_dir)

    lang_out_path = os.path.join(work_dir, 'OriginalStrings.csv')
    if args.lang_out_path:
        lang_out_path = os.path.abspath(args.lang_out_path)
    print('[ExcelTool] Excel key path:', lang_out_path)

    all_keys_out_path = None
    if args.all_keys_out_path:
        all_keys_out_path = os.path.abspath(args.all_keys_out_path)
        print('[ExcelTool] all keys output path:', all_keys_out_path)

    if args.m == 'data':
        convert_excel_to_data(work_dir, json_out_dir, cfg_fp, all_keys_out_path)
    elif args.m == 'language':
        generate_language_excel(work_dir, lang_cfg_fp, cfg_fp, lang_out_path)

    print('[ExcelTool] Done')

if __name__ == '__main__':
    main()
